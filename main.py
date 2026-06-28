import asyncio
import os
import sqlite3

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except Exception:
    psycopg2 = None
    RealDictCursor = None
from datetime import datetime, date
from pathlib import Path
from typing import Optional, Callable, Dict, Any, Awaitable

from aiogram import Bot, Dispatcher, F, Router, BaseMiddleware
from aiogram.filters import CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import CallbackQuery, FSInputFile, Message, ReplyKeyboardMarkup, KeyboardButton, TelegramObject
from aiogram.utils.keyboard import InlineKeyboardBuilder
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

load_dotenv(override=True)
BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
SELLER_LOGIN = os.getenv("SELLER_LOGIN", "sotuvchi")
SELLER_PASSWORD = os.getenv("SELLER_PASSWORD", "12345")
CONTROLLER_LOGIN = os.getenv("CONTROLLER_LOGIN", "nazoratchi")
CONTROLLER_PASSWORD = os.getenv("CONTROLLER_PASSWORD", "54321")
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
USE_POSTGRES = DATABASE_URL.startswith(("postgres://", "postgresql://"))
DB_PATH = Path(os.getenv("SQLITE_PATH", "bot.db"))
REPORT_DIR = Path("reports")
REPORT_DIR.mkdir(exist_ok=True)

FILIALS = [
    "1-Niyozbosh", "2-Xalqabod", "3-Gulbahor", "4-Kasblar", "5-Kids1",
    "6-Kids2", "7-Do’stobod", "8-Olmazor", "9-Chinoz", "10-Krasin",
    "11-Pitiletka", "12-Qo’rg’oncha", "13-Kids 3", "14-Oqqo’rg’on", "15-Qo’shyog’och"
]

router = Router()
USER_MENU_CAT = {}  # user_id -> house/stationery, menu tugmalari uchun

class CallbackNoLoadingMiddleware(BaseMiddleware):
    async def __call__(
        self,
        handler: Callable[[TelegramObject, Dict[str, Any]], Awaitable[Any]],
        event: TelegramObject,
        data: Dict[str, Any],
    ) -> Any:
        if isinstance(event, CallbackQuery):
            try:
                await event.answer()
            except Exception:
                pass
        return await handler(event, data)

router.callback_query.middleware(CallbackNoLoadingMiddleware())

class LoginState(StatesGroup):
    role = State()
    login = State()
    password = State()

class Form(StatesGroup):
    text = State()
    number = State()
    price_buy = State()
    price_sell = State()
    select_book = State()
    select_level = State()
    select_item = State()
    select_filial = State()
    qty = State()
    water_qty = State()
    water_price = State()
    money_accept = State()

# ---------- DB ----------
def _pg_sql(q: str) -> str:
    """SQLite uslubidagi ? placeholderlarni PostgreSQL uchun %s ga o'zgartiradi."""
    return q.replace("?", "%s")

def db():
    if USE_POSTGRES:
        if psycopg2 is None:
            raise RuntimeError("PostgreSQL uchun psycopg2-binary o'rnatilmagan")
        return psycopg2.connect(DATABASE_URL)
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con

def init_db():
    if USE_POSTGRES:
        with db() as con:
            with con.cursor() as cur:
                cur.execute("""
                CREATE TABLE IF NOT EXISTS books(
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL
                );
                CREATE TABLE IF NOT EXISTS levels(
                    id SERIAL PRIMARY KEY,
                    book_id INTEGER NOT NULL,
                    name TEXT NOT NULL,
                    buy_price DOUBLE PRECISION DEFAULT 0,
                    sell_price DOUBLE PRECISION DEFAULT 0,
                    UNIQUE(book_id,name)
                );
                CREATE TABLE IF NOT EXISTS book_stock(
                    level_id INTEGER PRIMARY KEY,
                    qty INTEGER DEFAULT 0
                );
                CREATE TABLE IF NOT EXISTS book_sales(
                    id SERIAL PRIMARY KEY,
                    dt TEXT,
                    buyer_type TEXT,
                    filial TEXT,
                    book_id INTEGER,
                    level_id INTEGER,
                    book_name TEXT,
                    level_name TEXT,
                    qty INTEGER,
                    buy_price DOUBLE PRECISION,
                    sell_price DOUBLE PRECISION,
                    total_buy DOUBLE PRECISION,
                    total_sell DOUBLE PRECISION,
                    profit DOUBLE PRECISION
                );
                CREATE TABLE IF NOT EXISTS items(
                    id SERIAL PRIMARY KEY,
                    category TEXT,
                    name TEXT,
                    buy_price DOUBLE PRECISION DEFAULT 0,
                    UNIQUE(category,name)
                );
                CREATE TABLE IF NOT EXISTS item_stock(
                    item_id INTEGER PRIMARY KEY,
                    qty INTEGER DEFAULT 0
                );
                CREATE TABLE IF NOT EXISTS item_gives(
                    id SERIAL PRIMARY KEY,
                    dt TEXT,
                    category TEXT,
                    filial TEXT,
                    item_id INTEGER,
                    item_name TEXT,
                    qty INTEGER,
                    buy_price DOUBLE PRECISION,
                    total_buy DOUBLE PRECISION
                );
                CREATE TABLE IF NOT EXISTS water_sales(
                    id SERIAL PRIMARY KEY,
                    dt TEXT,
                    filial TEXT,
                    qty INTEGER,
                    sell_price DOUBLE PRECISION,
                    total_sell DOUBLE PRECISION
                );
                CREATE TABLE IF NOT EXISTS payments(
                    id SERIAL PRIMARY KEY,
                    dt TEXT,
                    sales_total DOUBLE PRECISION,
                    accepted DOUBLE PRECISION,
                    debt_change DOUBLE PRECISION,
                    debt_after DOUBLE PRECISION,
                    note TEXT
                );
                """)
                cur.execute("ALTER TABLE book_sales ADD COLUMN IF NOT EXISTS book_name TEXT")
                cur.execute("ALTER TABLE book_sales ADD COLUMN IF NOT EXISTS level_name TEXT")
                cur.execute("ALTER TABLE item_gives ADD COLUMN IF NOT EXISTS item_name TEXT")
            con.commit()
        return

    with db() as con:
        con.executescript("""
        CREATE TABLE IF NOT EXISTS books(id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL);
        CREATE TABLE IF NOT EXISTS levels(id INTEGER PRIMARY KEY AUTOINCREMENT, book_id INTEGER NOT NULL, name TEXT NOT NULL,
            buy_price REAL DEFAULT 0, sell_price REAL DEFAULT 0, UNIQUE(book_id,name));
        CREATE TABLE IF NOT EXISTS book_stock(level_id INTEGER PRIMARY KEY, qty INTEGER DEFAULT 0);
        CREATE TABLE IF NOT EXISTS book_sales(id INTEGER PRIMARY KEY AUTOINCREMENT, dt TEXT, buyer_type TEXT, filial TEXT,
            book_id INTEGER, level_id INTEGER, book_name TEXT, level_name TEXT, qty INTEGER, buy_price REAL, sell_price REAL, total_buy REAL, total_sell REAL, profit REAL);
        CREATE TABLE IF NOT EXISTS items(id INTEGER PRIMARY KEY AUTOINCREMENT, category TEXT, name TEXT, buy_price REAL DEFAULT 0, UNIQUE(category,name));
        CREATE TABLE IF NOT EXISTS item_stock(item_id INTEGER PRIMARY KEY, qty INTEGER DEFAULT 0);
        CREATE TABLE IF NOT EXISTS item_gives(id INTEGER PRIMARY KEY AUTOINCREMENT, dt TEXT, category TEXT, filial TEXT,
            item_id INTEGER, item_name TEXT, qty INTEGER, buy_price REAL, total_buy REAL);
        CREATE TABLE IF NOT EXISTS water_sales(id INTEGER PRIMARY KEY AUTOINCREMENT, dt TEXT, filial TEXT, qty INTEGER, sell_price REAL, total_sell REAL);
        CREATE TABLE IF NOT EXISTS payments(id INTEGER PRIMARY KEY AUTOINCREMENT, dt TEXT, sales_total REAL, accepted REAL, debt_change REAL, debt_after REAL, note TEXT);
        """)
        for q in [
            "ALTER TABLE book_sales ADD COLUMN book_name TEXT",
            "ALTER TABLE book_sales ADD COLUMN level_name TEXT",
            "ALTER TABLE item_gives ADD COLUMN item_name TEXT",
        ]:
            try:
                con.execute(q)
                con.commit()
            except sqlite3.OperationalError:
                pass


def rows(q, args=()):
    if USE_POSTGRES:
        with db() as con:
            with con.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(_pg_sql(q), args)
                return cur.fetchall()
    with db() as con:
        return con.execute(q, args).fetchall()

def one(q,args=()):
    if USE_POSTGRES:
        with db() as con:
            with con.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(_pg_sql(q), args)
                return cur.fetchone()
    with db() as con:
        return con.execute(q,args).fetchone()

def execute(q,args=()):
    if USE_POSTGRES:
        with db() as con:
            with con.cursor() as cur:
                cur.execute(_pg_sql(q), args)
            con.commit()
        return None
    with db() as con:
        cur=con.execute(q,args); con.commit(); return cur.lastrowid

def _first_value(row):
    if row is None:
        return None
    if isinstance(row, dict):
        return next(iter(row.values()))
    return row[0]

def stock_qty(level_id):
    r=one("SELECT qty FROM book_stock WHERE level_id=?",(level_id,)); return int(_first_value(r) or 0) if r else 0

def item_qty(item_id):
    r=one("SELECT qty FROM item_stock WHERE item_id=?",(item_id,)); return int(_first_value(r) or 0) if r else 0

INTEGRITY_ERRORS = (sqlite3.IntegrityError,)
if psycopg2 is not None:
    INTEGRITY_ERRORS = (sqlite3.IntegrityError, psycopg2.IntegrityError)

def fmt_money(value):
    try:
        value = float(value or 0)
    except (TypeError, ValueError):
        value = 0
    return f"{value:,.0f}".replace(",", ".") + " so'm"

def parse_money(text):
    raw = (text or "").lower().replace("so'm", "").replace("sum", "").replace("сум", "")
    raw = raw.replace(" ", "").replace(".", "").replace(",", "")
    if not raw.isdigit():
        raise ValueError("summa noto'g'ri")
    return float(raw)

def now(): return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
def today_prefix(): return date.today().strftime("%Y-%m-%d")

def role_of(uid:int)->Optional[str]:
    # oddiy sessiya: user fsmda saqlanadi, bu funksiya kelajak uchun joy
    return None

def rows(q, args=()):
    with db() as con: return con.execute(q, args).fetchall()
def one(q,args=()):
    with db() as con: return con.execute(q,args).fetchone()
def execute(q,args=()):
    with db() as con:
        cur=con.execute(q,args); con.commit(); return cur.lastrowid

def stock_qty(level_id):
    r=one("SELECT qty FROM book_stock WHERE level_id=?",(level_id,)); return int(r[0]) if r else 0

def item_qty(item_id):
    r=one("SELECT qty FROM item_stock WHERE item_id=?",(item_id,)); return int(r[0]) if r else 0

# ---------- Keyboard ----------
def kb(buttons, cols=2):
    b=InlineKeyboardBuilder()
    for text, data in buttons: b.button(text=text, callback_data=data)
    b.adjust(cols)
    return b.as_markup()

def rkb(texts, cols=2):
    rows=[]
    for i in range(0, len(texts), cols):
        rows.append([KeyboardButton(text=t) for t in texts[i:i+cols]])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True, one_time_keyboard=False)

def main_kb(): return rkb(["🛒 Sotuvchi","👮 Nazoratchi"],2)
def seller_kb(): return rkb(["📚 Kitob sotuv","🧹 Xo'jalik maxsulotlari","✏️ Konstovar","💧 Suv berish","📊 Hisobot","⬅️ Chiqish"],2)
def controller_kb(): return rkb(["📦 Ombor","💳 Qarzni to‘lash","💰 Kunlik tushum","📩 Kunlik Excel","📅 Oylik Excel","🚪 Chiqish"],2)
def book_kb(): return rkb(["➕ Kitob yaratish","➕ Daraja yaratish","💰 Narx qo'yish","✏️ Narx tahrirlash","✏️ Kitob tahrirlash","✏️ Daraja tahrirlash","🗑 Kitob o'chirish","🗑 Daraja o'chirish","📥 Kitobni omborga qo'shish","🛍 Kitob sotish","📦 Kitoblar ombori","⬅️ Orqaga"],2)
def item_kb(cat):
    # Xo'jalik va Konstovar menyusi pastdagi yozish joyida chiqadi
    return rkb(["➕ Mahsulot qo'shish","✏️ Mahsulot tahrirlash","📦 Mahsulot ombori","📥 Mahsulot omborga qo'shish","💰 Mahsulot narx qo'yish","🚚 Filialga berish","⬅️ Orqaga"],2)

def filial_reply_kb():
    return rkb(FILIALS + ["⬅️ Orqaga"], 2)

def book_sale_kb():
    return rkb(["👨‍🎓 O'quvchiga sotish", "🏢 Filialga sotish", "⬅️ Orqaga"], 2)
def reports_kb(): return rkb(["📚 Kitob Excel","🧹 Xo'jalik Excel","✏️ Konstovar Excel","💧 Suv Excel","📊 Umumiy Excel","💳 Qarzdorlik Excel","💳 Qarzdorlik ko'rish","⬅️ Orqaga"],2)
def water_kb(): return rkb(["💧 Suv berish","⬅️ Orqaga"],2)

def book_buttons(prefix):
    bs=rows("SELECT * FROM books ORDER BY name")
    if not bs: return None
    return kb([(b['name'],f"{prefix}:{b['id']}") for b in bs]+[("⬅️ Orqaga","seller:books")],1)
def level_buttons(book_id,prefix):
    ls=rows("SELECT * FROM levels WHERE book_id=? ORDER BY name",(book_id,))
    if not ls: return None
    return kb([(l['name'],f"{prefix}:{l['id']}") for l in ls]+[("⬅️ Orqaga","seller:books")],1)
def item_buttons(cat,prefix):
    its=rows("SELECT * FROM items WHERE category=? ORDER BY name",(cat,))
    if not its: return None
    return kb([(i['name'],f"{prefix}:{i['id']}") for i in its]+[("⬅️ Orqaga",f"seller:{cat}")],1)
def filial_buttons(prefix): return kb([(f,f"{prefix}:{i}") for i,f in enumerate(FILIALS)]+[("⬅️ Orqaga","back:seller")],1)

# ---------- Excel ----------
def style_sheet(ws):
    fill=PatternFill("solid", fgColor="1F4E78"); font=Font(color="FFFFFF", bold=True)
    thin=Side(style="thin", color="CCCCCC")
    for c in ws[1]: c.fill=fill; c.font=font; c.alignment=Alignment(horizontal="center"); c.border=Border(top=thin,left=thin,right=thin,bottom=thin)
    for row in ws.iter_rows():
        for c in row: c.border=Border(top=thin,left=thin,right=thin,bottom=thin); c.alignment=Alignment(vertical="center")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=max(12,min(35,max(len(str(c.value or "")) for c in col)+2))

def save_report(kind):
    wb=Workbook(); ws=wb.active
    if kind=="books":
        ws.title="Kitob sotuv"; ws.append(["Sana","Kimga","Filial","Kitob","Daraja","Soni","Asil narx","Sotuv narx","Jami asil","Jami asil narx","Foyda"])
        data=rows("SELECT s.*, COALESCE(s.book_name,b.name,'O''chirilgan kitob') AS book, COALESCE(s.level_name,l.name,'O''chirilgan daraja') AS level FROM book_sales s LEFT JOIN books b ON b.id=s.book_id LEFT JOIN levels l ON l.id=s.level_id ORDER BY s.dt DESC")
        for r in data: ws.append([r['dt'],r['buyer_type'],r['filial'] or '',r['book'],r['level'],r['qty'],fmt_money(r['buy_price']),fmt_money(r['sell_price']),fmt_money(r['total_buy']),fmt_money(r['total_sell']),fmt_money(r['profit'])])
    elif kind in ("house","stationery"):
        ws.title=kind; ws.append(["Sana","Kategoriya","Filial","Mahsulot","Soni","Olingan narx","Jami qiymat"])
        data=rows("SELECT g.*, COALESCE(g.item_name,i.name,'O''chirilgan mahsulot') AS item FROM item_gives g LEFT JOIN items i ON i.id=g.item_id WHERE g.category=? ORDER BY g.dt DESC",(kind,))
        for r in data: ws.append([r['dt'],r['category'],r['filial'],r['item'],r['qty'],fmt_money(r['buy_price']),fmt_money(r['total_buy'])])
    elif kind=="water":
        ws.title="Suv"; ws.append(["Sana","Filial","Baklashka","Har bir asil narx","Jami asil narx"])
        for r in rows("SELECT * FROM water_sales ORDER BY dt DESC"): ws.append([r['dt'],r['filial'],r['qty'],fmt_money(r['sell_price']),fmt_money(r['total_sell'])])
    elif kind=="debt":
        ws.title="Qarzdorlik"
        ws.append(["Sana", "Bugungi sotuv", "Qabul qilingan", "Qarz o'zgarishi", "Jami qarz", "Izoh"])
        for r in rows("SELECT * FROM payments ORDER BY dt DESC"):
            ws.append([r['dt'], fmt_money(r['sales_total']), fmt_money(r['accepted']), fmt_money(r['debt_change']), fmt_money(r['debt_after']), r['note'] or ""])
    else:
        ws.title="Umumiy"; ws.append(["Bo'lim","Sana","Izoh","Soni","Asil qiymat","Sotuv summa","Foyda"])
        for r in rows("SELECT s.*, COALESCE(s.book_name,b.name,'O''chirilgan kitob') AS book, COALESCE(s.level_name,l.name,'O''chirilgan daraja') AS level FROM book_sales s LEFT JOIN books b ON b.id=s.book_id LEFT JOIN levels l ON l.id=s.level_id ORDER BY s.dt DESC"):
            ws.append(["Kitob",r['dt'],f"{r['book']} / {r['level']} / {r['buyer_type']} {r['filial'] or ''}",r['qty'],fmt_money(r['total_buy']),fmt_money(r['total_sell']),fmt_money(r['profit'])])
        for r in rows("SELECT g.*, COALESCE(g.item_name,i.name,'O''chirilgan mahsulot') AS item FROM item_gives g LEFT JOIN items i ON i.id=g.item_id ORDER BY g.dt DESC"):
            ws.append([r['category'],r['dt'],f"{r['item']} -> {r['filial']}",r['qty'],fmt_money(r['total_buy']),fmt_money(0),fmt_money(0)])
        for r in rows("SELECT * FROM water_sales ORDER BY dt DESC"):
            ws.append(["Suv",r['dt'],f"{r['filial']} ga berildi",r['qty'],fmt_money(r['total_sell']),fmt_money(0),fmt_money(0)])
    style_sheet(ws)
    path=REPORT_DIR/f"{kind}_hisobot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(path); return path

# ---------- Handlers ----------
@router.message(CommandStart())
async def start(m:Message, state:FSMContext):
    await state.clear(); await m.answer("Assalomu alaykum! Panelni tanlang:", reply_markup=main_kb())

@router.callback_query(F.data=="back:start")
async def back_start(c:CallbackQuery,state:FSMContext):
    await state.clear(); await c.message.answer("Panelni tanlang:", reply_markup=main_kb())
@router.callback_query(F.data=="back:seller")
async def back_seller(c:CallbackQuery,state:FSMContext):
    await state.clear(); await c.message.answer("Sotuvchi paneli:", reply_markup=seller_kb())

@router.callback_query(F.data.startswith("role:"))
async def role(c:CallbackQuery,state:FSMContext):
    r=c.data.split(':')[1]; await state.update_data(role=r); await state.set_state(LoginState.login)
    await c.message.edit_text("Login kiriting:")
@router.message(LoginState.login)
async def login(m:Message,state:FSMContext):
    await state.update_data(login=m.text.strip()); await state.set_state(LoginState.password); await m.answer("Parol kiriting:")
@router.message(LoginState.password)
async def password(m:Message,state:FSMContext):
    d=await state.get_data(); role=d['role']; login=d['login']; pwd=m.text.strip()
    ok=(role=='seller' and login==SELLER_LOGIN and pwd==SELLER_PASSWORD) or (role=='controller' and login==CONTROLLER_LOGIN and pwd==CONTROLLER_PASSWORD)
    if not ok: await state.clear(); await m.answer("❌ Login yoki parol xato.", reply_markup=main_kb()); return
    await state.clear(); await state.update_data(auth=role)
    await m.answer("✅ Kirildi.", reply_markup=seller_kb() if role=='seller' else controller_kb())


# ---------- Reply keyboard message handlers ----------
@router.message(F.text.in_({"🛒 Sotuvchi", "👮 Nazoratchi"}))
async def msg_choose_role(m: Message, state: FSMContext):
    role = "seller" if m.text == "🛒 Sotuvchi" else "controller"
    await state.update_data(role=role)
    await state.set_state(LoginState.login)
    await m.answer("Login kiriting:")

@router.message(F.text.in_({"⬅️ Chiqish", "🚪 Chiqish"}))
async def msg_exit(m: Message, state: FSMContext):
    await state.clear()
    await m.answer("Panelni tanlang:", reply_markup=main_kb())

@router.message(F.text == "⬅️ Orqaga")
async def msg_back_seller(m: Message, state: FSMContext):
    await state.clear()
    await m.answer("Sotuvchi paneli:", reply_markup=seller_kb())

@router.message(F.text == "📚 Kitob sotuv")
async def msg_seller_books(m: Message):
    await m.answer("📚 Kitob sotuv bo'limi:", reply_markup=book_kb())

@router.message(F.text == "🧹 Xo'jalik maxsulotlari")
async def msg_house(m: Message, state: FSMContext):
    USER_MENU_CAT[m.from_user.id] = 'house'
    await state.update_data(cat='house', menu_cat='house')
    await m.answer("🧹 Xo'jalik maxsulotlari:", reply_markup=item_kb('house'))

@router.message(F.text == "✏️ Konstovar")
async def msg_stationery(m: Message, state: FSMContext):
    USER_MENU_CAT[m.from_user.id] = 'stationery'
    await state.update_data(cat='stationery', menu_cat='stationery')
    await m.answer("✏️ Konstovar:", reply_markup=item_kb('stationery'))

@router.message(F.text == "💧 Suv berish")
async def msg_water_menu(m: Message, state: FSMContext):
    await state.update_data(action='water_filial')
    await m.answer("Qaysi filialga suv berilsin?", reply_markup=filial_reply_kb())

@router.message(F.text == "📊 Hisobot")
async def msg_reports_menu(m: Message):
    await m.answer("📊 Hisobotlar:", reply_markup=reports_kb())

@router.message(F.text == "➕ Kitob yaratish")
async def msg_book_create(m: Message, state: FSMContext):
    await state.update_data(action='book_create')
    await state.set_state(Form.text)
    await m.answer("Kitob nomini yozing:")

@router.message(F.text == "➕ Daraja yaratish")
async def msg_level_create(m: Message):
    markup = book_buttons('levelbook')
    await m.answer("Qaysi kitobga daraja yaratmoqchisiz?" if markup else "Avval kitob yarating.", reply_markup=markup or book_kb())

@router.message(F.text.in_({"💰 Narx qo'yish", "✏️ Narx tahrirlash", "📥 Kitobni omborga qo'shish", "✏️ Kitob tahrirlash", "🗑 Kitob o'chirish"}))
async def msg_book_actions(m: Message, state: FSMContext):
    mp={"💰 Narx qo'yish":"price:set", "✏️ Narx tahrirlash":"price:edit", "📥 Kitobni omborga qo'shish":"stock:addbook", "✏️ Kitob tahrirlash":"book:edit", "🗑 Kitob o'chirish":"book:delete"}
    action=mp[m.text]
    await state.update_data(action=action)
    markup=book_buttons('choosebook')
    await m.answer("Qaysi kitobni tanlaysiz?" if markup else "Avval kitob yarating.", reply_markup=markup or book_kb())

@router.message(F.text == "✏️ Daraja tahrirlash")
async def msg_level_edit(m: Message):
    markup=book_buttons('editlevelbook')
    await m.answer("Qaysi kitob darajasini tahrirlaysiz?" if markup else "Avval kitob yarating.", reply_markup=markup or book_kb())

@router.message(F.text == "🗑 Daraja o'chirish")
async def msg_level_delete(m: Message):
    markup=book_buttons('deletelevelbook')
    await m.answer("Qaysi kitob darajasini o'chirasiz?" if markup else "Avval kitob yarating.", reply_markup=markup or book_kb())

@router.message(F.text == "📦 Kitoblar ombori")
async def msg_stock_books(m: Message):
    data=rows("SELECT b.name book,l.name level,COALESCE(s.qty,0) qty FROM levels l JOIN books b ON b.id=l.book_id LEFT JOIN book_stock s ON s.level_id=l.id ORDER BY b.name,l.name")
    txt="📦 Kitoblar ombori:\n\n"+"\n".join([f"{r['book']} / {r['level']}: {r['qty']} ta" for r in data]) if data else "Ombor bo'sh."
    await m.answer(txt, reply_markup=book_kb())

@router.message(F.text == "🛍 Kitob sotish")
async def msg_sale_bookmenu(m: Message):
    await m.answer("Kimga sotiladi?", reply_markup=book_sale_kb())

@router.message(F.text == "💧 Suv berish")
async def msg_water_sell(m: Message, state: FSMContext):
    await state.update_data(action='water_filial')
    await m.answer("Qaysi filialga suv berilsin?", reply_markup=filial_reply_kb())


@router.message(F.text == "👨‍🎓 O'quvchiga sotish")
async def msg_sale_student_reply(m: Message, state: FSMContext):
    await state.update_data(action='sale_book', sale_type='student')
    await m.answer("Qaysi kitob?", reply_markup=book_buttons('salebook') or book_kb())

@router.message(F.text == "🏢 Filialga sotish")
async def msg_sale_branch_reply(m: Message, state: FSMContext):
    await state.update_data(action='sale_branch_filial', sale_type='branch')
    await m.answer("Qaysi filialga?", reply_markup=filial_reply_kb())

@router.message(F.text.in_({"➕ Mahsulot qo'shish", "✏️ Mahsulot tahrirlash", "📦 Mahsulot ombori", "📥 Mahsulot omborga qo'shish", "💰 Mahsulot narx qo'yish", "🚚 Filialga berish"}))
async def msg_item_menu_actions(m: Message, state: FSMContext):
    d = await state.get_data()
    cat = d.get('cat') or d.get('menu_cat') or USER_MENU_CAT.get(m.from_user.id)
    if cat in ('house', 'stationery'):
        USER_MENU_CAT[m.from_user.id] = cat
        await state.update_data(cat=cat, menu_cat=cat)
    if cat not in ('house', 'stationery'):
        await m.answer("Avval Xo'jalik maxsulotlari yoki Konstovar bo'limiga kiring.", reply_markup=seller_kb())
        return
    if m.text == "➕ Mahsulot qo'shish":
        await state.update_data(action='item_create', cat=cat, menu_cat=cat)
        await state.set_state(Form.text)
        await m.answer("Mahsulot nomini kiriting:")
    elif m.text == "✏️ Mahsulot tahrirlash":
        await state.update_data(cat=cat, menu_cat=cat)
        await m.answer("Qaysi mahsulot tahrirlanadi?", reply_markup=item_buttons(cat,'itemedit') or item_kb(cat))
    elif m.text == "💰 Mahsulot narx qo'yish":
        await state.update_data(cat=cat, menu_cat=cat)
        await m.answer("Qaysi mahsulotga narx qo'yiladi?", reply_markup=item_buttons(cat,'itemprice') or item_kb(cat))
    elif m.text == "📥 Mahsulot omborga qo'shish":
        await state.update_data(cat=cat, menu_cat=cat)
        await m.answer("Qaysi mahsulot omborga qo'shiladi?", reply_markup=item_buttons(cat,'itemstockadd') or item_kb(cat))
    elif m.text == "🚚 Filialga berish":
        await state.update_data(action='item_give_filial', cat=cat, menu_cat=cat, give_item=True)
        await m.answer("Qaysi filialga?", reply_markup=filial_reply_kb())
    elif m.text == "📦 Mahsulot ombori":
        data=rows("SELECT i.name,COALESCE(s.qty,0) qty,i.buy_price FROM items i LEFT JOIN item_stock s ON s.item_id=i.id WHERE i.category=? ORDER BY i.name",(cat,))
        txt="📦 Ombor:\n\n"+("\n".join([f"• {r['name']} — {r['qty']} ta | narx {fmt_money(r['buy_price'])}" for r in data]) or "Bo'sh")
        await m.answer(txt, reply_markup=item_kb(cat))

@router.message(F.text.in_(set(FILIALS)))
async def msg_filial_reply(m: Message, state: FSMContext):
    d = await state.get_data()
    action = d.get('action')
    if action == 'water_filial':
        await state.update_data(filial=m.text)
        await state.set_state(Form.water_qty)
        await m.answer("Nechta baklashka berildi?")
    elif action == 'sale_branch_filial':
        await state.update_data(filial=m.text, action='sale_book', sale_type='branch')
        await m.answer("Qaysi kitob?", reply_markup=book_buttons('salebook') or book_kb())
    elif action == 'item_give_filial':
        cat = d.get('cat') or d.get('menu_cat')
        await state.update_data(filial=m.text, cat=cat, menu_cat=cat, give_item=True)
        await m.answer("Qaysi mahsulot?", reply_markup=item_buttons(cat,'itemgivepick') or item_kb(cat))
    else:
        await m.answer("Bu filial tanlash joyi emas.")

@router.message(F.text.in_({"📚 Kitob Excel","🧹 Xo'jalik Excel","✏️ Konstovar Excel","💧 Suv Excel","📊 Umumiy Excel","💳 Qarzdorlik Excel"}))
async def msg_reports_excel_reply(m: Message):
    mp={"📚 Kitob Excel":"books","🧹 Xo'jalik Excel":"house","✏️ Konstovar Excel":"stationery","💧 Suv Excel":"water","📊 Umumiy Excel":"all","💳 Qarzdorlik Excel":"debt"}
    path = save_report(mp[m.text])
    await m.answer_document(FSInputFile(path), caption="✅ Hisobot tayyor.", reply_markup=reports_kb())

@router.message(F.text == "💳 Qarzdorlik ko'rish")
async def msg_debt_view(m: Message):
    await show_debt_message(m, reports_kb())

@router.message(F.text.in_({"📦 Ombor", "💰 Kunlik tushum", "📩 Kunlik Excel", "📅 Oylik Excel", "💳 Qarzni to‘lash"}))
async def msg_controller_buttons(m: Message, state: FSMContext):
    if m.text == "📦 Ombor":
        await ctrl_stock_message(m)
    elif m.text == "💰 Kunlik tushum":
        await ctrl_daily_message(m)
    elif m.text == "📩 Kunlik Excel":
        path=save_report('all')
        await m.answer_document(FSInputFile(path), caption="✅ Kunlik Excel tayyor.", reply_markup=controller_kb())
    elif m.text == "📅 Oylik Excel":
        path=save_report('all')
        await m.answer_document(FSInputFile(path), caption="✅ Oylik Excel tayyor.", reply_markup=controller_kb())
    elif m.text == "💳 Qarzni to‘lash":
        await start_payment_message(m, state)

@router.callback_query(F.data=="seller:books")
async def seller_books(c:CallbackQuery): await c.message.answer("📚 Kitob sotuv bo'limi:", reply_markup=book_kb())
@router.callback_query(F.data=="seller:house")
async def house(c:CallbackQuery, state: FSMContext):
    USER_MENU_CAT[c.from_user.id]='house'
    await state.update_data(cat='house', menu_cat='house')
    await c.message.answer("🧹 Xo'jalik maxsulotlari:", reply_markup=item_kb('house'))
@router.callback_query(F.data=="seller:stationery")
async def stationery(c:CallbackQuery, state: FSMContext):
    USER_MENU_CAT[c.from_user.id]='stationery'
    await state.update_data(cat='stationery', menu_cat='stationery')
    await c.message.answer("✏️ Konstovar:", reply_markup=item_kb('stationery'))
@router.callback_query(F.data=="seller:reports")
async def reports(c:CallbackQuery): await c.message.answer("📊 Hisobotlar:", reply_markup=reports_kb())

# Books CRUD
@router.callback_query(F.data=="book:create")
async def book_create(c,state): await state.update_data(action='book_create'); await state.set_state(Form.text); await c.message.edit_text("Kitob nomini yozing:")
@router.message(Form.text)
async def text_input(m,state):
    d=await state.get_data(); action=d.get('action'); text=m.text.strip(); keep_cat=None
    try:
        if action=='book_create':
            execute("INSERT INTO books(name) VALUES(?)",(text,)); await m.answer("✅ Kitob yaratildi. Endi bu kitobga daraja biriktiring: Daraja yaratish bo'limidan qilasiz.", reply_markup=book_kb())
        elif action=='level_name':
            execute("INSERT INTO levels(book_id,name) VALUES(?,?)",(d['book_id'],text)); await m.answer("✅ Daraja yaratildi.", reply_markup=book_kb())
        elif action=='book_rename':
            execute("UPDATE books SET name=? WHERE id=?",(text,d['book_id'])); await m.answer("✅ Kitob nomi tahrirlandi.", reply_markup=book_kb())
        elif action=='level_rename':
            execute("UPDATE levels SET name=? WHERE id=?",(text,d['level_id'])); await m.answer("✅ Daraja nomi tahrirlandi.", reply_markup=book_kb())
        elif action in ('item_create','item_rename'):
            cat=d['cat']; keep_cat=cat; USER_MENU_CAT[m.from_user.id]=cat
            if action=='item_create': execute("INSERT INTO items(category,name) VALUES(?,?)",(cat,text)); msg="✅ Mahsulot qo'shildi."
            else: execute("UPDATE items SET name=? WHERE id=?",(text,d['item_id'])); msg="✅ Mahsulot tahrirlandi."
            await m.answer(msg, reply_markup=item_kb(cat))
        else: await m.answer("Noma'lum amal. /start")
    except INTEGRITY_ERRORS:
        await m.answer("❌ Bu nom oldin yaratilgan.")
    await state.clear()
    if keep_cat in ('house', 'stationery'):
        await state.update_data(cat=keep_cat, menu_cat=keep_cat)

@router.callback_query(F.data=="level:create")
async def level_create(c,state):
    markup=book_buttons('levelbook')
    await c.message.answer("Qaysi kitobga daraja yaratmoqchisiz?" if markup else "Avval kitob yarating.", reply_markup=markup or book_kb())
@router.callback_query(F.data.startswith("levelbook:"))
async def level_book(c,state):
    await state.update_data(action='level_name',book_id=int(c.data.split(':')[1])); await state.set_state(Form.text); await c.message.edit_text("Daraja nomini yozing:")

@router.callback_query(F.data.in_({"price:set","price:edit","stock:addbook","book:edit","book:delete"}))
async def choose_book(c,state):
    action=c.data; await state.update_data(action=action)
    markup=book_buttons('choosebook')
    await c.message.answer("Qaysi kitob?" if markup else "Kitob yo'q.", reply_markup=markup or book_kb())
@router.callback_query(F.data.startswith("choosebook:"))
async def book_chosen(c,state):
    d=await state.get_data(); book_id=int(c.data.split(':')[1]); action=d['action']; await state.update_data(book_id=book_id)
    if action=='book:edit': await state.update_data(action='book_rename'); await state.set_state(Form.text); await c.message.edit_text("Yangi kitob nomini yozing:"); return
    if action=='book:delete': execute("DELETE FROM books WHERE id=?",(book_id,)); execute("DELETE FROM levels WHERE book_id=?",(book_id,)); await c.message.answer("✅ Kitob va unga tegishli darajalar o'chirildi.", reply_markup=book_kb()); return
    markup=level_buttons(book_id,'chooselevel')
    await c.message.answer("Qaysi daraja?" if markup else "Bu kitobda daraja yo'q.", reply_markup=markup or book_kb())
@router.callback_query(F.data.startswith("chooselevel:"))
async def level_chosen(c,state):
    d=await state.get_data(); lid=int(c.data.split(':')[1]); await state.update_data(level_id=lid)
    if d['action'] in ('price:set','price:edit'):
        await state.set_state(Form.price_buy); await c.message.edit_text("Qanchaga harid qildingiz?")
    elif d['action']=='stock:addbook':
        await state.update_data(stock_type='book'); await state.set_state(Form.qty); await c.message.edit_text("Nechta qo'shish kerak?")

@router.message(Form.price_buy)
async def buy_price(m,state):
    try: val=parse_money(m.text);
    except: await m.answer("Faqat summa yozing."); return
    d=await state.get_data()
    if d.get('action')=='item_price':
        execute("UPDATE items SET buy_price=? WHERE id=?",(val,d['item_id']))
        cat=d['cat']; USER_MENU_CAT[m.from_user.id]=cat
        await state.clear()
        await state.update_data(cat=cat, menu_cat=cat)
        await m.answer("✅ Mahsulot narxi saqlandi.", reply_markup=item_kb(cat))
        return
    await state.update_data(buy_price=val); await state.set_state(Form.price_sell); await m.answer("Qanchaga sotiladi?")
@router.message(Form.price_sell)
async def sell_price(m,state):
    try: val=parse_money(m.text);
    except: await m.answer("Faqat summa yozing."); return
    d=await state.get_data(); execute("UPDATE levels SET buy_price=?, sell_price=? WHERE id=?",(d['buy_price'],val,d['level_id']))
    await state.clear(); await m.answer("✅ Narx saqlandi.", reply_markup=book_kb())

@router.callback_query(F.data=="level:edit")
async def ledit(c,state): await state.update_data(action='level_edit_pickbook'); await level_create_like(c,'editlevel')
@router.callback_query(F.data=="level:delete")
async def ldel(c,state): await state.update_data(action='level_delete_pickbook'); await level_create_like(c,'deletelevel')
async def level_create_like(c,prefix):
    markup=book_buttons(prefix+'book'); await c.message.answer("Qaysi kitob?" if markup else "Kitob yo'q.", reply_markup=markup or book_kb())
@router.callback_query(F.data.startswith("editlevelbook:"))
async def edit_level_book(c,state):
    markup=level_buttons(int(c.data.split(':')[1]),'editlevel'); await c.message.answer("Qaysi daraja?", reply_markup=markup or book_kb())
@router.callback_query(F.data.startswith("deletelevelbook:"))
async def del_level_book(c,state):
    markup=level_buttons(int(c.data.split(':')[1]),'deletelevel'); await c.message.answer("Qaysi daraja o'chiriladi?", reply_markup=markup or book_kb())
@router.callback_query(F.data.startswith("editlevel:"))
async def edit_level(c,state): await state.update_data(action='level_rename',level_id=int(c.data.split(':')[1])); await state.set_state(Form.text); await c.message.edit_text("Yangi daraja nomini yozing:")
@router.callback_query(F.data.startswith("deletelevel:"))
async def del_level(c,state): execute("DELETE FROM levels WHERE id=?",(int(c.data.split(':')[1]),)); await c.message.answer("✅ Daraja o'chirildi.", reply_markup=book_kb())

@router.message(Form.qty)
async def qty_input(m,state):
    try: qty=int(m.text.strip())
    except: await m.answer("Faqat son yozing."); return
    d=await state.get_data()
    if d.get('stock_type')=='book':
        execute("INSERT INTO book_stock(level_id,qty) VALUES(?,?) ON CONFLICT(level_id) DO UPDATE SET qty=qty+excluded.qty",(d['level_id'],qty))
        await state.clear(); await m.answer("✅ Kitob omborga qo'shildi.", reply_markup=book_kb())
    elif d.get('stock_type')=='item':
        execute("INSERT INTO item_stock(item_id,qty) VALUES(?,?) ON CONFLICT(item_id) DO UPDATE SET qty=qty+excluded.qty",(d['item_id'],qty))
        cat=d['cat']; USER_MENU_CAT[m.from_user.id]=cat; await state.clear(); await state.update_data(cat=cat, menu_cat=cat); await m.answer("✅ Omborga qo'shildi.", reply_markup=item_kb(cat))
    elif d.get('sale_type') in ('student','branch'):
        lid=d['level_id']; left=stock_qty(lid)
        if left<qty: await m.answer(f"❌ Omborda yetarli emas. Qolgan: {left} ta"); return
        lev=one("SELECT l.*, b.name AS book_name FROM levels l JOIN books b ON b.id=l.book_id WHERE l.id=?",(lid,)); book_id=lev['book_id']; filial=d.get('filial')
        total_buy=lev['buy_price']*qty; total_sell=lev['sell_price']*qty
        execute("INSERT INTO book_sales(dt,buyer_type,filial,book_id,level_id,book_name,level_name,qty,buy_price,sell_price,total_buy,total_sell,profit) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (now(), 'Oquvchi' if d['sale_type']=='student' else 'Filial', filial, book_id,lid,lev['book_name'],lev['name'],qty,lev['buy_price'],lev['sell_price'],total_buy,total_sell,total_sell-total_buy))
        execute("UPDATE book_stock SET qty=qty-? WHERE level_id=?",(qty,lid))
        await state.clear(); await m.answer(f"✅ Kitob sotildi. Jami: {fmt_money(total_sell)}", reply_markup=book_kb())
    elif d.get('give_item'):
        iid=d['item_id']; left=item_qty(iid)
        if left<qty: await m.answer(f"❌ Omborda yetarli emas. Qolgan: {left} ta"); return
        it=one("SELECT * FROM items WHERE id=?",(iid,)); total=it['buy_price']*qty
        execute("INSERT INTO item_gives(dt,category,filial,item_id,item_name,qty,buy_price,total_buy) VALUES(?,?,?,?,?,?,?,?)",(now(),d['cat'],d['filial'],iid,it['name'],qty,it['buy_price'],total))
        execute("UPDATE item_stock SET qty=qty-? WHERE item_id=?",(qty,iid))
        cat=d['cat']; USER_MENU_CAT[m.from_user.id]=cat; await state.clear(); await state.update_data(cat=cat, menu_cat=cat); await m.answer("✅ Filialga berildi va ombordan ayirildi.", reply_markup=item_kb(cat))

@router.callback_query(F.data=="stock:books")
async def stock_books(c):
    data=rows("SELECT b.name book,l.name level,COALESCE(s.qty,0) qty,l.buy_price,l.sell_price FROM levels l JOIN books b ON b.id=l.book_id LEFT JOIN book_stock s ON s.level_id=l.id ORDER BY b.name,l.name")
    txt="📦 Kitoblar ombori:\n\n"+("\n".join([f"📚 {r['book']} / {r['level']} — {r['qty']} ta | olish {fmt_money(r['buy_price'])} | sotish {fmt_money(r['sell_price'])}" for r in data]) or "Bo'sh")
    await c.message.answer(txt, reply_markup=book_kb())

@router.callback_query(F.data=="sale:bookmenu")
async def sale_menu(c): await c.message.answer("Kitob kimga sotiladi?", reply_markup=book_sale_kb())
@router.callback_query(F.data=="sale:student")
async def sale_student(c,state): await state.update_data(action='sale_book',sale_type='student'); await c.message.answer("Qaysi kitob?", reply_markup=book_buttons('salebook') or book_kb())
@router.callback_query(F.data=="sale:branch")
async def sale_branch(c,state): await state.update_data(action='sale_book',sale_type='branch'); await c.message.edit_text("Qaysi filialga?", reply_markup=filial_buttons('salefilial'))
@router.callback_query(F.data.startswith("salefilial:"))
async def sale_filial(c,state): await state.update_data(filial=FILIALS[int(c.data.split(':')[1])]); await c.message.answer("Qaysi kitob?", reply_markup=book_buttons('salebook') or book_kb())
@router.callback_query(F.data.startswith("salebook:"))
async def sale_book(c,state):
    bid=int(c.data.split(':')[1]); await state.update_data(book_id=bid); await c.message.answer("Qaysi daraja?", reply_markup=level_buttons(bid,'salelevel') or book_kb())
@router.callback_query(F.data.startswith("salelevel:"))
async def sale_level(c,state):
    lid=int(c.data.split(':')[1]); left=stock_qty(lid); await state.update_data(level_id=lid); await state.set_state(Form.qty)
    if left<=0: await state.clear(); await c.message.answer("❌ Bu kitob omborda qolmagan.", reply_markup=book_kb()); return
    await c.message.edit_text(f"Omborda {left} ta bor. Nechta sotildi?")

# Items
@router.callback_query(F.data.startswith("item:create:"))
async def item_create(c,state): cat=c.data.split(':')[2]; USER_MENU_CAT[c.from_user.id]=cat; await state.update_data(action='item_create',cat=cat,menu_cat=cat); await state.set_state(Form.text); await c.message.edit_text("Mahsulot nomini kiriting:")
@router.callback_query(F.data.startswith("item:edit:"))
async def item_edit(c,state): cat=c.data.split(':')[2]; USER_MENU_CAT[c.from_user.id]=cat; await state.update_data(cat=cat,menu_cat=cat); await c.message.answer("Qaysi mahsulot tahrirlanadi?", reply_markup=item_buttons(cat,'itemedit') or item_kb(cat))
@router.callback_query(F.data.startswith("itemedit:"))
async def item_edit_pick(c,state): await state.update_data(action='item_rename',item_id=int(c.data.split(':')[1])); await state.set_state(Form.text); await c.message.edit_text("Yangi nomini yozing:")
@router.callback_query(F.data.startswith("item:price:"))
async def item_price(c,state): cat=c.data.split(':')[2]; USER_MENU_CAT[c.from_user.id]=cat; await state.update_data(cat=cat,menu_cat=cat); await c.message.answer("Qaysi mahsulotga narx qo'yiladi?", reply_markup=item_buttons(cat,'itemprice') or item_kb(cat))
@router.callback_query(F.data.startswith("itemprice:"))
async def item_price_pick(c,state):
    await state.update_data(action='item_price', item_id=int(c.data.split(':')[1]))
    await state.set_state(Form.price_buy)
    await c.message.edit_text("Qanchaga olindi?")

@router.callback_query(F.data.startswith("item:addstock:"))
async def item_addstock(c,state): cat=c.data.split(':')[2]; USER_MENU_CAT[c.from_user.id]=cat; await state.update_data(cat=cat,menu_cat=cat); await c.message.answer("Qaysi mahsulot omborga qo'shiladi?", reply_markup=item_buttons(cat,'itemstockadd') or item_kb(cat))
@router.callback_query(F.data.startswith("itemstockadd:"))
async def item_stock_add_pick(c,state): await state.update_data(stock_type='item',item_id=int(c.data.split(':')[1])); await state.set_state(Form.qty); await c.message.edit_text("Nechta keldi?")
@router.callback_query(F.data.startswith("item:give:"))
async def item_give(c,state): cat=c.data.split(':')[2]; USER_MENU_CAT[c.from_user.id]=cat; await state.update_data(cat=cat,menu_cat=cat,give_item=True); await c.message.edit_text("Qaysi filialga?", reply_markup=filial_buttons('itemfilial'))
@router.callback_query(F.data.startswith("itemfilial:"))
async def item_filial(c,state): d=await state.get_data(); await state.update_data(filial=FILIALS[int(c.data.split(':')[1])]); await c.message.answer("Qaysi mahsulot?", reply_markup=item_buttons(d['cat'],'itemgivepick') or item_kb(d['cat']))
@router.callback_query(F.data.startswith("itemgivepick:"))
async def item_give_pick(c,state): await state.update_data(item_id=int(c.data.split(':')[1])); await state.set_state(Form.qty); await c.message.edit_text("Nechta beriladi?")
@router.callback_query(F.data.startswith("item:stock:"))
async def item_stock(c):
    cat=c.data.split(':')[2]; data=rows("SELECT i.name,COALESCE(s.qty,0) qty,i.buy_price FROM items i LEFT JOIN item_stock s ON s.item_id=i.id WHERE i.category=? ORDER BY i.name",(cat,))
    txt="📦 Ombor:\n\n"+("\n".join([f"• {r['name']} — {r['qty']} ta | narx {fmt_money(r['buy_price'])}" for r in data]) or "Bo'sh")
    await c.message.answer(txt, reply_markup=item_kb(cat))

# item price custom via generic buy_price conflict workaround: use state data no level_id
# Override buy_price handler by editing existing function would be complex, patched below by checking item_id in buy_price above? It doesn't. Need patch file later.

# Water
@router.callback_query(F.data=="seller:water")
async def water(c): await c.message.answer("💧 Suv berish:", reply_markup=water_kb())
@router.callback_query(F.data=="water:sell")
async def water_sell(c,state): await state.update_data(action='water_filial'); await c.message.answer("Qaysi filialga suv berilsin?", reply_markup=filial_reply_kb())
@router.callback_query(F.data.startswith("waterfilial:"))
async def water_filial(c,state): await state.update_data(filial=FILIALS[int(c.data.split(':')[1])]); await state.set_state(Form.water_qty); await c.message.edit_text("Nechta baklashka berildi?")
@router.message(Form.water_qty)
async def water_qty(m,state):
    try: q=int(m.text.strip())
    except: await m.answer("Faqat son yozing."); return
    await state.update_data(qty=q); await state.set_state(Form.water_price); await m.answer("Har bir baklashka qanchadan olingan?")
@router.message(Form.water_price)
async def water_price(m,state):
    try: p=parse_money(m.text)
    except: await m.answer("Faqat summa yozing."); return
    d=await state.get_data(); total=d['qty']*p
    execute("INSERT INTO water_sales(dt,filial,qty,sell_price,total_sell) VALUES(?,?,?,?,?)",(now(),d['filial'],d['qty'],p,total))
    await state.clear(); await m.answer(f"✅ Suv berish saqlandi. Jami asil qiymat: {fmt_money(total)}", reply_markup=seller_kb())

# Reports and controller
@router.callback_query(F.data.startswith("report:"))
async def report(c):
    kind=c.data.split(':')[1]; path=save_report(kind); await c.message.answer_document(FSInputFile(path), caption="✅ Excel hisobot tayyor.")
@router.callback_query(F.data=="ctrl:daily")
async def ctrl_daily(c): await c.message.answer_document(FSInputFile(save_report('all')), caption="📅 Kunlik/umumiy Excel hisobot")
@router.callback_query(F.data=="ctrl:monthly")
async def ctrl_monthly(c): await c.message.answer_document(FSInputFile(save_report('all')), caption="🗓 Oylik/umumiy Excel hisobot")
@router.callback_query(F.data=="ctrl:stock")
async def ctrl_stock(c):
    await c.message.answer(controller_stock_text(), reply_markup=controller_kb())

def today_sales_total():
    # Suv filialga beriladi, sotuv hisoblanmaydi. Qarzdorlik va tushumga faqat haqiqiy sotuvlar kiradi.
    tp=today_prefix()
    b=one("SELECT COALESCE(SUM(total_sell),0) s FROM book_sales WHERE dt LIKE ?",(tp+'%',))['s']
    return float(b or 0)
def current_debt():
    r=one("SELECT debt_after FROM payments ORDER BY id DESC LIMIT 1"); return float(r['debt_after']) if r else 0.0
@router.callback_query(F.data=="ctrl:payment")
async def payment(c,state):
    total=today_sales_total(); debt=current_debt(); await state.update_data(sales_total=total,old_debt=debt); await state.set_state(Form.money_accept)
    await c.message.edit_text(f"💵 Bugungi sotuv: {fmt_money(total)}\nOldingi qarzdorlik: {fmt_money(debt)}\n\nQancha pul qabul qilindi?")
@router.message(Form.money_accept)
async def money_accept(m,state):
    try: accepted=parse_money(m.text)
    except: await m.answer("Faqat summa yozing."); return
    d=await state.get_data(); debt_after=max(0, d['old_debt'] + d['sales_total'] - accepted); change=debt_after-d['old_debt']
    execute("INSERT INTO payments(dt,sales_total,accepted,debt_change,debt_after,note) VALUES(?,?,?,?,?,?)",(now(),d['sales_total'],accepted,change,debt_after,'Pul qabul qilindi'))
    
    if change > 0:
        msg = f"✅ Saqlandi. {fmt_money(change)} qarzdorlikka qo'shildi.\nHozirgi qarzdorlik: {fmt_money(debt_after)}"
    elif change < 0:
        msg = f"✅ Saqlandi. Ortiqcha pul qarzdan ayirildi: {fmt_money(abs(change))}.\nHozirgi qarzdorlik: {fmt_money(debt_after)}"
    else:
        msg = f"✅ Saqlandi. Qarz o'zgarmadi.\nHozirgi qarzdorlik: {fmt_money(debt_after)}"
    await state.clear(); await m.answer(msg, reply_markup=controller_kb())
def debt_text():
    ps=rows("SELECT * FROM payments ORDER BY dt DESC LIMIT 30")
    if not ps:
        return "💳 Qarzdorlik tarixi:\n\nHali yo'q"
    lines=[]
    for p in ps:
        change = float(p['debt_change'] or 0)
        if change > 0:
            holat = f"➕ qarz qo'shildi: {fmt_money(change)}"
        elif change < 0:
            holat = f"➖ qarzdan yechildi: {fmt_money(abs(change))}"
        else:
            holat = "✅ qarz o'zgarmadi"
        lines.append(f"{p['dt']} | sotuv {fmt_money(p['sales_total'])} | olindi {fmt_money(p['accepted'])} | {holat} | jami qarz {fmt_money(p['debt_after'])}")
    return "💳 Qarzdorlik tarixi:\n\n" + "\n".join(lines)

@router.callback_query(F.data=="seller:debt")
async def seller_debt(c):
    await c.message.answer(debt_text(), reply_markup=reports_kb())

@router.callback_query(F.data=="ctrl:debt")
async def debt(c):
    await c.message.answer(debt_text(), reply_markup=controller_kb())

async def show_debt_message(m: Message, markup=None):
    await m.answer(debt_text(), reply_markup=markup or reports_kb())

def controller_stock_text():
    # Nazoratchi omborida kitob, xo'jalik mahsulotlari va konstovar umumiy qoldiqlari ko'rsatiladi.
    book_data = rows("""
        SELECT b.name book, l.name level, COALESCE(s.qty,0) qty, l.buy_price, l.sell_price
        FROM levels l
        JOIN books b ON b.id=l.book_id
        LEFT JOIN book_stock s ON s.level_id=l.id
        ORDER BY b.name,l.name
    """)
    house_data = rows("""
        SELECT i.name, COALESCE(s.qty,0) qty, i.buy_price
        FROM items i
        LEFT JOIN item_stock s ON s.item_id=i.id
        WHERE i.category='house'
        ORDER BY i.name
    """)
    stationery_data = rows("""
        SELECT i.name, COALESCE(s.qty,0) qty, i.buy_price
        FROM items i
        LEFT JOIN item_stock s ON s.item_id=i.id
        WHERE i.category='stationery'
        ORDER BY i.name
    """)

    parts = ["📦 Umumiy ombor"]

    parts.append("\n📚 Kitoblar:")
    if book_data:
        parts.extend([
            f"• {r['book']} / {r['level']} — {r['qty']} ta | olish {fmt_money(r['buy_price'])} | sotish {fmt_money(r['sell_price'])}"
            for r in book_data
        ])
    else:
        parts.append("Bo'sh")

    parts.append("\n🧹 Xo'jalik mahsulotlari:")
    if house_data:
        parts.extend([f"• {r['name']} — {r['qty']} ta | asil narx {fmt_money(r['buy_price'])}" for r in house_data])
    else:
        parts.append("Bo'sh")

    parts.append("\n✏️ Konstovar:")
    if stationery_data:
        parts.extend([f"• {r['name']} — {r['qty']} ta | asil narx {fmt_money(r['buy_price'])}" for r in stationery_data])
    else:
        parts.append("Bo'sh")

    return "\n".join(parts)

async def ctrl_stock_message(m: Message):
    await m.answer(controller_stock_text(), reply_markup=controller_kb())

async def ctrl_daily_message(m: Message):
    total=today_sales_total()
    debt=current_debt()
    await m.answer(f"💰 Bugungi sotuv: {fmt_money(total)}\n💳 Hozirgi qarzdorlik: {fmt_money(debt)}", reply_markup=controller_kb())

async def start_payment_message(m: Message, state: FSMContext):
    total=today_sales_total()
    debt=current_debt()
    await state.update_data(sales_total=total, old_debt=debt)
    await state.set_state(Form.money_accept)
    await m.answer(f"💵 Bugungi sotuv: {fmt_money(total)}\nOldingi qarzdorlik: {fmt_money(debt)}\n\nQancha pul qabul qilindi?")

async def main():
    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN .env ichida yozilmagan")
    init_db()
    bot=Bot(BOT_TOKEN)
    dp=Dispatcher(storage=MemoryStorage())
    dp.include_router(router)
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
